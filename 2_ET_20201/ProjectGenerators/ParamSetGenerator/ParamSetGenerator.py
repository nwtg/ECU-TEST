import os

from openpyxl import load_workbook

import tt

from tts.core.project.generator.ParamGenerator import ParamGenerator
from tts.lib.common.files.PathHandler import MakeAbsoluteWorkspacePath # Toto
from tts.core.report.ReportConstants import COLUMN_NAME


class ParamSetGenerator(ParamGenerator):
    """
    This is a small example on how a parameter generator should be implemented.
    """

    ID = '56fe8f8f-e11d-11df-a48c-12386baaabec'
    NAME = 'Excel example generator'
    DESCRIPTION = 'Read cells from excel'
    SERIALIZE = {'filename': ('FILENAME', 'string', r'')}
    COLUMN_NAME = 1
    COLUMN_DRIVER_REQUEST = 2
    COLUMN_ENGINE_SPEED = 3

    """
    Here the instance variables are declared for serialization.
    """

    def __init__(self):
        """
        Method, to initalize the instance variables.
        """
        self.filename = r''
        """
        Initialize the name of the file.
        """
        self.workbook = None
        self.workSheetName = "Data"
        self.paramSheet = None

    def GetFilename(self):
        """
        Returns name of csv file.
        For use in the configuration dialog.
        :return: Name of csv file.
        :rtype: str
        """
        return self.filename

    def SetFilename(self, filename):
        """
        Sets name of csv file.
        For use in the configuration dialog.
        :param filename: Name of csv file.
        :type filename: str
        """
        self.filename = filename
    PFilename = property(GetFilename, SetFilename)

    def GetDialog(self):
        """
        Returns the configuration dialog of the parameter generator.
        The dialog needs a reference to the parameter generator for getting and setting
        the parameter generator data.
        :rtype: L{DlgFileSelect}
        :return: wxDialog
        """
        from .DlgFileSelect import DlgFileSelect
        return DlgFileSelect(None, self)

    def PreGeneration(self):
        """
        A file descriptor to the excel file will be opend.
        """
        filePath = self.__ValidatePath(self.filename)

        self.workbook = load_workbook(filePath)
        self.paramSheet = self.workbook[self.workSheetName]

    def GenerationIterator(self):
        """
        Method, to implement the iteration over the lines in the excel file.
        """
        for row in range(self.paramSheet.max_row):
            nameOfRow = str(self.paramSheet.cell(row=row + 1, column=self.COLUMN_NAME).value)
            driverRequest = self.paramSheet.cell(row=row + 1,
                                                 column=self.COLUMN_DRIVER_REQUEST).value
            engineSpeed = self.paramSheet.cell(row=row + 1, column=self.COLUMN_ENGINE_SPEED).value

            cycleData = self.CreateCycleData(name="[{name}] DriverRequest = {request} %".format(
                                                                            name=nameOfRow,
                                                                            request=driverRequest))
            cycleData.AddParameter('DriverRequest', int(driverRequest))
            cycleData.AddParameter('EngineSpeed', int(engineSpeed))

            # # also possible
            # params = {'firstNumber':int(firstNumber), 'secondNumber':int(secondNumber)}
            # cycleData = self.CreateCycleData(paramSet=params, name="%s (%s)" % (nameOfRow, row))

            yield cycleData

    def Check(self):
        """
        Is executed when the containing project is checked.
        Verify the existence of the file.
        """
        errors = []
        try:
            self.__ValidatePath(self.filename)
        except Exception as e:
            errors.append(str(e))

        return errors

    def __ValidatePath(self, filePath):
        """
        Checks the referenced file path. Returns an absolute path, if the referenced file exist.

        :param filePath: absolute or relative to the packages directory.
        :type filePath: str

        :return: absolute file path
        :rtype: str

        :raise tt.Error: if the referenced file path is invalid.
        """
        if not filePath:
            raise tt.Error('Undefined excel file path!')

        filePath = MakeAbsoluteWorkspacePath(filePath) # Toto
        if not os.path.isfile(filePath):
            raise tt.Error('"%s" is not a valid file path!' % filePath)

        return filePath
