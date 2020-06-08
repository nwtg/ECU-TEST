# -*- coding: utf-8 -*-

'''
Report generator settings:

  'templateFile' - defines the path to the Excel template. Defined path can be absolute or 
                   relative to the 'report.py' containing directory.                   
  
  'reportName'   - defines the name for the generated Excel report. Defined name will be 
                   extended with a time prefix to avoid Excel problems with already opened 
                   files.


To configure the Excel report generator to match the requirements of a modified report template 
following settings are available:

  'execResult'   - defines a position for the report result. Should be a tuple containing row and 
                   column indexes e.g. "(1,1)" or "-" if the corresponding value should not
                   be reported.
                 
  'execDuration' - defines a position for the test execution duration (in hours). Should be a tuple 
                   containing row and column indexes e.g. "(1,1)" or "-" if the corresponding value  
                   should not be reported.
                             
  'execDate'     - defines a position for the test execution date. Should be a tuple containing row 
                   and column indexes e.g. "(1,1)" or "-" if the corresponding value should not
                   be reported.
                
  'author'       - defines a position for the tester name. Should be a tuple containing row and 
                   column indexes e.g. "(1,1)" or "-" if the corresponding value should not
                   be reported.
             
  'testCaseResultStartRow'    - defines the start row index for test case property entries e.g. "1"    
  
  'testCaseNamesColumn'       - defines the column index for test case names e.g. "1"
  
  'testCaseResultColumn'      - defines the column index for test case results e.g. "2"
  
  'testCaseDescriptionColumn' - defines the column index for test descriptions e.g. "3" 
  
  'attrCaption'      - defines the position of the attribute caption

  'attrStartCoord'   - defines the start position of the attribute table

  'colorResultCells' - if "True" the background of the result cells will be colored depending on 
                       the result state. To avoid these use "False" or "-" 
  
  'tableFrames' - tuple containing the min and max column indexes and a boolean value e.g. 
                  "(1,5,True)" or "-". If defined, the generated test case result entries will 
                  be surrounded by frames. 
                  Start and end rows depend on the 'testCaseResultStartRow' definition and the 
                  count of the generated entries. 
                  If the boolean value is set to "True" the start row for the frame generation 
                  will be decremented by one. These should be used, if a title row defined in the 
                  report template before execution.
  
  'configEntriesStartAt' - defines the start position for the configuration entries on the page. 
                           Should be a tuple containing row and column indexes e.g. "(1,1)" 
'''


import os
import shutil
import sys
import time

from ast import literal_eval as frStr
import tt
import win32gui
import win32con
from openpyxl.styles import Border, PatternFill, Alignment, Side, Font
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from datetime import datetime



REPORT_INFO_SETTINGS = {'execResult':'GetResult', 'execDuration':'GetDuration',
                        'execDate':'GetExecutionTime', 'author':'GetAuthor'}
RESULT_COLORS = {'INCONCLUSIVE':'fff15c', 'NONE':'CDC5B9', 'ERROR':'B40000', 'FAILED':'F25757',
                'SUCCESS':'77FF5B'}
# Separator for connecting element identifier
PACKAGE_NAME_SEPARATOR = '\\\n'  # '\\'

def ProcessReport(reportName, reportApi):
    pkgRes = config = []
    attrs = None
    firstCycle = True
    try:
        for item in reportApi.IterItems():
            itemType = item.__class__.__name__
            if itemType == 'Project':
                if firstCycle and reportApi.IsProjectReport() and item.HasAttributes():
                    attrs = item.IterAttributes()
                pkgRes = _ProcessContainerElement(reportApi, item)
            if itemType == 'Package':
                pkgRes = [_ProcessPackage(item)]
            if itemType == 'Configuration':
                config = _ParseConfiguration(item)
    except Exception:
        exc = sys.exc_info()
        msg = u'Parsing the report file failed, caused by:\n%s' % exc[1]
        __raiseExceptionShowDialogBefore(msg, exc[0], exc[2])

    _CreateExecReportFromTemplate(reportApi, reportName, config, pkgRes, attrs)

def _ProcessContainerElement(reportApi, prjElem, prefix=u''):
    pkgRes = []
    for item in prjElem.IterItems():
        itemType = item.__class__.__name__
        if itemType == 'Package':
            pkgRes.append(_ProcessPackage(item, prefix))
        elif itemType in ['Project', 'ProjectElement']:
            # Ignore skipped project steps
            if item.IsSkipped():
                continue
            newPrefix = u'%s%s' % ((u'%s%s' % (prefix, PACKAGE_NAME_SEPARATOR))
                                 if prefix else u'', item.GetName())
            pkgRes.extend(_ProcessContainerElement(reportApi, item, newPrefix))
    return pkgRes

def _ProcessPackage(item, prefix=None):
    baseName = item.GetPrjCompName()
    packageName = u'%s%s' % ((u'%s%s' % (prefix, PACKAGE_NAME_SEPARATOR)) if prefix
                                                                          else u'', baseName)
    # If the package is skipped due to user decision (filer, commented out)
    # then the result verdict is manipulated here
    if item.IsSkipped():
        packageResult = u'(SKIPPED)'
        packageDescription = item.GetComment()
		# Toto
        duration = item.GetDuration()
    else:
        packageResult = item.GetResult()
        packageDescription = item.GetDescription()
		# Toto
        duration = item.GetDuration()
    return (packageName, packageResult, packageDescription, duration)

def __raiseExceptionShowDialogBefore(excMessage, excType=None, excTraceback=None):

    excMessage = u'Excel report generation failed, caused by:\n%s' % tt.Str(excMessage)

    win32gui.MessageBox(0, excMessage, u'Exception occurred!', win32con.MB_ICONERROR)

    if excType and excTraceback:
        raise excType, excMessage, excTraceback
    else:
        raise RuntimeError(excMessage)

def _ExtractTCFData(tcf):
    tcfData = {}

    # common
    editor = tcf.GetEditor()

    testBench = tcf.GetTeststand()

    dataDir = tcf.GetDataDir()

    pkgDir = tcf.GetPkgDir()

    tcfData['common'] = {'editor':editor, 'testBench':testBench, 'dataDir':dataDir,
                         'pkgDir':pkgDir}

    # modell
    modelDataList = []
    if tcf.HasModelConfigurations():
        for mdl in tcf.IterModelConfigurations():
            mdlId = mdl.GetTcfModelId()
            mdlPath = mdl.GetModel()
            timeBase = bool(mdl.GetTimebase())
            modelDataList.append({'id':mdlId, 'path':mdlPath, 'timeBase':timeBase})
    tcfData['models'] = modelDataList
    # Function definitions
#     if tcf.HasFunctionConfigurations():
#         for func in tcf.IterFunctionConfigurations():
#             func.GetTcfFctId()
#             pass

    # FIUs
    fiuDataList = []
    if tcf.HasEfsConfigurations():
        for efs in tcf.IterEfsConfigurations():
            fiuId = efs.GetTcfEfsId()
            fiuPath = efs.GetDb()
            fiuDataList.append({'id':fiuId, 'path':fiuPath})
    tcfData['fius'] = fiuDataList

    # ECUs
    ecuDataList = []
    if tcf.HasEcuConfigurations():
        for ecu in tcf.IterEcuConfigurations():
            ecuId = ecu.GetTcfEcuId()
            ecuSGBD = ecu.GetSgbd()
            ecuA2L = ecu.GetA2lFile()
            ecuHex = ecu.GetHexFile()
            ecuDataList.append({'id':ecuId, 'sgbd':ecuSGBD, 'a2l':ecuA2L, 'hex':ecuHex})
    tcfData['ecus'] = ecuDataList

    # Bus
    busDataList = []
    if tcf.HasBusConfigurations():
        for bus in tcf.IterBusConfigurations():
            busId = bus.GetTcfBusId()
            busDb = bus.GetDbPath()
            busCh = bus.GetFbxChn()
            busDataList.append({'id':busId, 'db':busDb, 'channel':busCh})
    tcfData['buses'] = busDataList

    # Global Constants
    constFileList = []
    for constFile in tcf.IterConstFiles():
        constFileList.append(constFile)
    tcfData['constantFiles'] = constFileList

    return tcfData

def _ExtractTBCData(tbc):
    tbcData = {}

    # Tools
    toolDataList = []
    if tbc.HasTools():
        for tool in tbc.IterTools():
            toolId = tool.GetTbcToolId()
            toolHost = tool.GetLocation()
            toolStatus = tool.GetStatus()
            toolVersion = tool.GetVersion()
            if not toolVersion:
                toolVersion = u'unknown'
            toolDataList.append({'id': toolId, 'host':toolHost, 'status':toolStatus,
                                 'version':toolVersion})
    tbcData['tools'] = toolDataList

    return tbcData

def _ParseConfiguration(cfg):
    configData = {}

    # TCF
    tcf = cfg.GetTestConfiguration()
    tcfData = None
    if tcf:
        tcfData = _ExtractTCFData(tcf)
        tcfPath = tcf.GetPath()
        configData['tcfPath'] = tcfPath

    # TBC
    tbc = cfg.GetTestBenchConfiguration()
    tbcData = None
    if tbc:
        tbcData = _ExtractTBCData(tbc)
        tbcPath = tbc.GetPath()
        configData['tbcPath'] = tbcPath

    configData['tcf'] = tcfData
    configData['tbc'] = tbcData
    return configData

def __GetSettingValue(reportApi, settingName):
    value = reportApi.GetSetting(settingName)
    if value.lower() in ['-', 'n/a', 'none']:
        return None
    return frStr(value)

def _CreateTableBorder(sheet, cell_range, outerSide, innerSide):
    top = Border(top=outerSide, bottom=innerSide, left=innerSide, right=innerSide)
    left = Border(top=innerSide, bottom=innerSide, left=outerSide, right=innerSide)
    right = Border(top=innerSide, bottom=innerSide, left=innerSide, right=outerSide)
    bottom = Border(top=innerSide, bottom=outerSide, left=innerSide, right=innerSide)
    rows = list(sheet[cell_range])


    for row in rows[1:-1]:
        for cell in row[1:-1]:
            cell.border = Border(top=innerSide, bottom=innerSide, left=innerSide, right=innerSide)

    for cell in rows[0]:
        cell.border = top
    for cell in rows[-1]:
        cell.border = bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = left
        r.border = right

    rows[0][0].border = Border(top=outerSide, bottom=innerSide, left=outerSide, right=innerSide)
    rows[-1][0].border = Border(top=innerSide, bottom=outerSide, left=outerSide, right=innerSide)
    rows[0][-1].border = Border(top=outerSide, bottom=innerSide, left=innerSide, right=outerSide)
    rows[-1][-1].border = Border(top=innerSide, bottom=outerSide, left=innerSide, right=outerSide)

def _InsertText(sheet, text, row, col, font=None, fill=None, alignment=None):
    textCell = sheet.cell(row=row, column=col)
    textCell.font = font if font else Font()
    textCell.fill = fill if fill else PatternFill()
    textCell.alignment = alignment if alignment else Alignment()
    textCell.value = text

def _WritePkgName(sheet, reportApi, pkgName, row):
    namesCol = __GetSettingValue(reportApi, 'testCaseNamesColumn')
    _InsertText(sheet, pkgName, row, namesCol, alignment=Alignment(vertical='center'))

def _WritePkgResult(sheet, reportApi, res, row):
    resCol = __GetSettingValue(reportApi, 'testCaseResultColumn')
    coloredResBG = __GetSettingValue(reportApi, 'colorResultCells')

    if coloredResBG:
        fill = PatternFill(start_color=RESULT_COLORS.get(res),
                                   end_color=RESULT_COLORS.get(res), fill_type='solid')
    else:
        fill = None
    _InsertText(sheet, res, row, resCol, font=Font(b=True),
                alignment=Alignment(horizontal='center', vertical='center'),
                fill=fill)

def _WritePkgDescr(sheet, reportApi, descr, row):
    descrCol = __GetSettingValue(reportApi, 'testCaseDescriptionColumn')
    _InsertText(sheet, descr, row, descrCol, alignment=Alignment(vertical='center'))

# Toto
def _WritePkgPath(sheet, reportApi, path, row):
    pathCol = __GetSettingValue(reportApi, 'testCaseDurationColumn')
    _InsertText(sheet, path, row, pathCol, alignment=Alignment(vertical='center'))
    

def _CreateTestCaseResultBorder(sheet, startCell, endCell):
    thinBorder = Side(style="thin", color="000000")
    thickBorder = Side(style='thick', color="000000")
    _CreateTableBorder(sheet, u"{}:{}".format(startCell.coordinate, endCell.coordinate),
                      thickBorder, thinBorder)

def _WritePkgDetails(workbook, reportApi, pkgRes):
    startRow = __GetSettingValue(reportApi, 'testCaseResultStartRow')
    sheet = workbook.get_sheet_by_name('Overview')

    # Toto
    for i, (pkg, res, descr, duration) in enumerate(pkgRes):
        _WritePkgName(sheet, reportApi, pkg, startRow + i)
        _WritePkgResult(sheet, reportApi, res, startRow + i)
        _WritePkgDescr(sheet, reportApi, descr, startRow + i)
        # Toto
        _WritePkgPath(sheet, reportApi, duration, startRow + i)


    if pkgRes and __GetSettingValue(reportApi, 'tableFrames'):
        try:
            startCol, endCol, includeTitle = __GetSettingValue(reportApi, 'tableFrames')
            startRow = startRow - (1 if includeTitle else 0)
            startCol = int(startCol)
            endCol = int(endCol)
            endRow = startRow + len(pkgRes) - (0 if includeTitle else 1)
            _CreateTestCaseResultBorder(sheet,
                                sheet.cell(row=startRow, column=startCol),
                                sheet.cell(row=endRow, column=endCol))
        except Exception as e:
            raise RuntimeError(u'Error making Border: %s' % e)

def _WriteSummary(workbook, reportApi):
    coloredResBG = __GetSettingValue(reportApi, 'colorResultCells')
    reportInfo = reportApi.GetInfo()
    sheet = workbook.get_sheet_by_name('Overview')
    for setting in REPORT_INFO_SETTINGS:
        defined = __GetSettingValue(reportApi, setting)
        if defined is None:
            continue
        try:
            font = Font(color="808080")
            fill = None
            row, col = int(defined[0]), int(defined[1])
            text = getattr(reportInfo, REPORT_INFO_SETTINGS[setting])()
            # format:
            if setting == u'execDuration':
                text = time.gmtime(round(text))
                text = time.strftime('%H:%M:%S', text)
            elif setting == u'execDate':
                # FIXED : "execDate" is now an object of type datetime
                if isinstance(text, datetime):
                    text = text.strftime('%Y-%m-%d %H:%M:%S')
                else: 
                    text = text.rsplit('.', 1)[0]
            elif setting == u'execResult' and coloredResBG:
                fill = PatternFill(start_color=RESULT_COLORS.get(text),
                                   end_color=RESULT_COLORS.get(text), fill_type='solid')
                font = Font(b=True)

            _InsertText(sheet, text, row, col, font, fill)
        except Exception as e:
            raise RuntimeError(u'Setting parameter \'%s\' is incorrect! Error: %s'
                               % (setting, e))
            
    
def _WriteAttributes(workbook, reportApi, attrs):
    sheet = workbook.get_sheet_by_name('Overview')
    captionRow, captionCol = __GetSettingValue(reportApi, 'attrCaption')
    _InsertText(sheet, 'Attributes', captionRow, captionCol, Font(size=14, bold=True))

    startRow, startCol = __GetSettingValue(reportApi, 'attrStartCoord')

    for i, attr in enumerate(attrs):
        _InsertText(sheet, attr.GetName(), startRow + (i % 4), startCol + ((i / 4) * 2), Font(bold=True))
        _InsertText(sheet, attr.GetValue(), startRow + (i % 4), startCol + ((i / 4) * 2) + 1)


def _GetWorkbook(reportApi, reportName):

    # create Excel report from template
    reportDir = reportApi.GetReportDir()
    if not os.path.isdir(reportDir):
        raise RuntimeError(u'Report directory \'%s\' does not exist!' % reportDir)
    templateFile = reportApi.GetSetting('templateFile')
    if not templateFile:
        raise RuntimeError(u'No Excel template file referenced!')
    if not os.path.isabs(templateFile):
        curDir = os.path.dirname(__file__)
        templateFile = os.path.abspath(os.path.join(curDir, templateFile))
    if not os.path.isfile(templateFile):
        raise RuntimeError(u'Referenced Excel template \'%s\' does not exist!' % templateFile)

    exclReportName = reportApi.GetSetting('reportName')
    if not exclReportName:
        exclReportName = reportName
    name, ext = os.path.splitext(exclReportName)
    if not ext:
        ext = '.xlsx'
    exclReportName = '%s_%s%s' % (int(round(time.time())), "{}".format(name) , ext)
    exclReportPath = os.path.abspath(os.path.join(reportDir, exclReportName))

    if os.path.isfile(exclReportPath):
        os.remove(exclReportPath)
    shutil.copyfile(templateFile, exclReportPath)

    wb = load_workbook(exclReportPath)

    ws = _EnsureSheetExists(wb, u'Overview')
    ws.add_image(Image(os.path.abspath(os.path.join(curDir, 'Logo.jpg'))), 'B2')
    ws = _EnsureSheetExists(wb, u'Configuration')
    ws.add_image(Image(os.path.abspath(os.path.join(curDir, 'Logo.jpg'))), 'B2')
    return wb, exclReportPath



def _EnsureSheetExists(workbook, worksheetName):
    if worksheetName not in workbook.sheetnames:
        workbook.create_sheet(title=worksheetName)
    return workbook.get_sheet_by_name(worksheetName)


def _GetConfigStartPosition(reportApi):
    startcfgRow = cfgCol = None
    try:
        coords = __GetSettingValue(reportApi, 'configEntriesStartAt')
        if coords:
            startcfgRow = int(coords[0])
            cfgCol = int(coords[1])
    except:
        raise RuntimeError(u'Setting parameter \'configEntriesStartAt\' is incorrect! '
                           u'If defined, its value should be a tuple containing row and '
                           u'column indexes.')
    return startcfgRow, cfgCol


def _WriteConstFiles(sheet, startRow, startCol, constFiles):
    _InsertText(sheet, u'Global constants definition files', startRow, startCol, font=Font(bold=True))
    writtenRows = 1

    if len(constFiles) > 0:
        for constFile in constFiles:
            _InsertText(sheet, u'  %s' % constFile, startRow + writtenRows, startCol)
            writtenRows += 1
    else:
        _InsertText(sheet, u'  No global constants definition files configured',
                    startRow + writtenRows, startCol)
        writtenRows += 1

    return writtenRows


def _WriteBuses(sheet, startRow, startCol, buses):
    _InsertText(sheet, u'Bus access', startRow, startCol, font=Font(bold=True))
    writtenRows = 1
    if buses:
        keyList = [e['id'] for e in buses]
        keyList.reverse()
        buses = sorted(buses, key=lambda y:keyList.index(y['id']))
        for bus in buses:
            busId = bus['id']
            db = bus['db']
            ch = bus['channel']

            _InsertText(sheet, u'  %s' % busId, startRow + writtenRows, startCol)
            _InsertKeyValue(sheet, u'    Fibex channel', ch, startRow + writtenRows + 1, startCol)
            _InsertKeyValue(sheet, u'    Bus database', db, startRow + writtenRows + 2, startCol)
            writtenRows += 3

    else:
        _InsertText(sheet, u'  No Bus access configured', startRow + writtenRows, startCol)
        writtenRows += 1

    return writtenRows


def _WriteEcus(sheet, startRow, startCol, ecus):
    _InsertText(sheet, u'ECUs', startRow, startCol, font=Font(bold=True))
    writtenRows = 1
    if ecus:
        keyList = [e['id'] for e in ecus]
        keyList.reverse()
        ecus = sorted(ecus, key=lambda y:keyList.index(y['id']))
        for ecu in ecus:
            ecuId = ecu['id']
            sgbd = ecu['sgbd']
            a2l = ecu['a2l']
            hexFile = ecu['hex']

            _InsertText(sheet, u'  %s' % ecuId, startRow + writtenRows, startCol)
            _InsertKeyValue(sheet, u'    A2L file', a2l, startRow + writtenRows + 1, startCol)
            _InsertKeyValue(sheet, u'    Hex file', hexFile, startRow + writtenRows + 2, startCol)
            _InsertKeyValue(sheet, u'    SGBD', sgbd, startRow + writtenRows + 3, startCol)
            writtenRows += 4
    else:
        _InsertText(sheet, u'  No ECUs configured', startRow + writtenRows, startCol)
        writtenRows += 1

    return writtenRows

def _WriteFius(sheet, startRow, startCol, fius):
    _InsertText(sheet, u'Failure simulation', startRow, startCol, font=Font(bold=True))
    writtenRows = 1
    if fius:
        keyList = [e['id'] for e in fius]
        keyList.reverse()
        fius = sorted(fius, key=lambda y:keyList.index(y['id']))
        for fiu in fius:
            FiuId = fiu['id']
            path = fiu['path']

            _InsertText(sheet, u'  %s' % FiuId, startRow + writtenRows, startCol)
            _InsertKeyValue(sheet, u'    Fehlersimulations-Datei', path, startRow + writtenRows + 1, startCol)
            writtenRows += 2

    else:
        _InsertText(sheet, u'  No FIUs configured', startRow + writtenRows, startCol)
        writtenRows += 1

    return writtenRows



def _WriteModels(sheet, startRow, startCol, models):
    _InsertText(sheet, u'Models', startRow, startCol, font=Font(bold=True))
    writtenRows = 1
    if models:
        keyList = [e['id'] for e in models]
        keyList.reverse()
        mdls = sorted(models, key=lambda y:keyList.index(y['id']))
        for mdl in mdls:
            modelId = mdl['id']
            path = mdl['path']
            tb = mdl['timeBase']

            _InsertText(sheet, u'  %s' % modelId, startRow + writtenRows, startCol)
            _InsertKeyValue(sheet, u'    Model', path, startRow + writtenRows + 1, startCol)
            _InsertKeyValue(sheet, u'    Us as time base', tb, startRow + writtenRows + 2 , startCol)
            writtenRows += 3
    else:
        _InsertText(sheet, u'  No Models configured', startRow + writtenRows, startCol)
        writtenRows += 1

    return writtenRows

def _InsertKeyValue(sheet, key, value, startRow, startCol):
    _InsertText(sheet, key, startRow, startCol)
    _InsertText(sheet, value, startRow, startCol + 1)


def _WriteCommonData(sheet, startRow, startCol, commonData):
    _InsertText(sheet, u'Common', startRow, startCol, font=Font(bold=True))

    _InsertKeyValue(sheet, u'  Name of tester', commonData['editor'], startRow + 1, startCol)
    _InsertKeyValue(sheet, u'  Name of test computer', commonData['testBench'], startRow + 2, startCol)
    _InsertKeyValue(sheet, u'  Data directory', commonData['dataDir'], startRow + 3, startCol)
    _InsertKeyValue(sheet, u'  Package directory', commonData['pkgDir'], startRow + 4, startCol)

    return 5

def _WriteTcfDetails(sheet, tcfData, tcfPath, startRow, startCol):
    _InsertText(sheet, 'Used configuration file: %s' % tcfPath, startRow, startCol,
                font=Font(color='808080'))
    writtenRows = 1
    writtenRows += _WriteCommonData(sheet, startRow + writtenRows, startCol,
                                    tcfData['common'])

    writtenRows += _WriteModels(sheet, startRow + writtenRows, startCol,
                                    tcfData['models'])

    writtenRows += _WriteFius(sheet, startRow + writtenRows, startCol,
                                    tcfData['fius'])

    writtenRows += _WriteEcus(sheet, startRow + writtenRows, startCol,
                                    tcfData['ecus'])

    writtenRows += _WriteBuses(sheet, startRow + writtenRows, startCol,
                                    tcfData['buses'])

    writtenRows += _WriteConstFiles(sheet, startRow + writtenRows, startCol,
                                    tcfData['constantFiles'])
    return writtenRows


def _WriteTbcDetails(sheet, tbcData, tbcPath, startRow, startCol):
    _InsertText(sheet, 'Used configuration file: %s' % tbcPath, startRow, startCol,
                font=Font(color='808080'))
    writtenRows = 1
    tools = tbcData['tools']
    if tools:
        keyList = [e['id'] for e in tools]
        keyList.reverse()
        tools = sorted(tools, key=lambda y:keyList.index(y['id']))

        for tool in tools:
            toolId = tool['id']
            host = tool['host']
            version = tool['version']
            if not version:
                version = _(u'unbekannt')
            statNames = {u'ON': u'active',
                         u'OFF': u'inactive',
                         u'ERR': u'erroneous'
                        }
            status = statNames.get(tool['status'], u'unknown')

            _InsertText(sheet, toolId, startRow + writtenRows, startCol, font=Font(bold=True))
            _InsertKeyValue(sheet, u'  Host', host, startRow + writtenRows + 1 , startCol)
            _InsertKeyValue(sheet, u'  Status', status, startRow + writtenRows + 2 , startCol)
            _InsertKeyValue(sheet, u'  Version', version, startRow + writtenRows + 3 , startCol)
            writtenRows += 4

    else:
        _InsertText(sheet, u'No tools configured', startRow, startCol)
        writtenRows += 1

    return writtenRows

def _WriteConfig(workbook, reportApi, config):
    startRow, startCol = _GetConfigStartPosition(reportApi)
    if startRow is not None and startCol is not None:
        # fill the configuration page
        sheet = _EnsureSheetExists(workbook, u'Configuration')

        _InsertText(sheet, u'Test configuration', startRow, startCol,
                    font=Font(size=14, bold=True))
        writtenRows = 2  # Eine Leerzeile
        tcfData = config['tcf']
        if tcfData:
            writtenRows += _WriteTcfDetails(sheet, tcfData, config['tcfPath'],
                                     startRow + writtenRows, startCol)
        else:
            _InsertText(sheet, u'No test configuration loaded', startRow, startCol,
                        font=Font(bold=True))
            writtenRows += 1

        startRow += (writtenRows + 2)  # Leerzeilen
        _InsertText(sheet, u'Test-bench configuration', startRow, startCol,
                    font=Font(size=14, bold=True))
        writtenRows = 2  # Eine Leerzeile
        tbcData = config['tbc']
        if tbcData:
            writtenRows += _WriteTbcDetails(sheet, tbcData, config['tbcPath'],
                                     startRow + writtenRows, startCol)
        else:
            _InsertText(sheet, u'No test-bench configuration loaded', startRow, startCol,
                        font=Font(bold=True))
            writtenRows += 1

def _CreateExecReportFromTemplate(reportApi, reportName, config, pkgRes, attrs):

    try:
        workbook, exclReportPath = _GetWorkbook(reportApi, reportName)
        _WriteSummary(workbook, reportApi)
        if attrs:
            _WriteAttributes(workbook, reportApi, attrs)
        _WritePkgDetails(workbook, reportApi, pkgRes)

        _WriteConfig(workbook, reportApi, config)

    except Exception:
        exc = sys.exc_info()
        msg = u'Writing the Excel report file failed, caused by:\n%s' % exc[1]
        __raiseExceptionShowDialogBefore(msg, exc[0], exc[2])
    finally:
        workbook.save(exclReportPath)
